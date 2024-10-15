import pandas as pd 
import openpyxl
import json
import math


f = open('selected_rooms.json')
selectedHalls = json.loads(f.read())
selectedRollNo = []

wb = openpyxl.load_workbook('2022_Name List - Deptwise1.xlsx') 
res = len(wb.sheetnames)
print(res)

all_sheets = []

for i in range(0,res):
    df = pd.read_excel('2022_Name List - Deptwise1.xlsx',usecols='C',sheet_name=i)
    for _, row in df.iterrows():
      reg = row
      selectedRollNo.append(reg.item())
output = {}

wb = openpyxl.load_workbook('2023_Name List.xlsx') 
res = len(wb.sheetnames)
print(res)

for i in range(0,res):
    df = pd.read_excel('2023_Name List.xlsx',usecols='C',sheet_name=i)
    for _, row in df.iterrows():
      reg = row
      selectedRollNo.append(reg.item())

sheets_dict = pd.read_excel('2022_Name List - Deptwise1.xlsx')
sheets_dict1 = pd.read_excel('2023_Name List.xlsx')

t = []
  
for i in selectedRollNo:
    if i == 'nan':
        continue
    else:
        key = str(i)[2:5]
        t.append(key)

t =set(t)

dict_dept = {"year 1": {}, "year 2": {}, "year 3": {}}

# Loop through each department identifier in 't'
for dept_id in t:
    if dept_id == 'nan':
        continue
    
    # Create a temporary list to hold roll numbers for each department
    temp_list = []
    
    # Collect roll numbers associated with the current department identifier
    for roll_no in selectedRollNo:
        if roll_no == 'nan':
            continue
        # Match department codes (e.g., UCF, BAC) from the roll numbers
        if dept_id == str(roll_no)[2:5]:
            temp_list.append(roll_no)
    
    # Sort the collected roll numbers for the current department
    temp_list.sort()
    
    # Organize sorted roll numbers into year categories
    for roll_no in temp_list:
        prefix = str(roll_no)[0:2]
        department_code = str(roll_no)[2:5]  # Extract department code from characters 3 to 5
        if prefix == "22":
            dict_dept["year 3"].setdefault(department_code, []).append(roll_no)
        elif prefix == "23":
            dict_dept["year 2"].setdefault(department_code, []).append(roll_no)
        elif prefix == "24":
            dict_dept["year 1"].setdefault(department_code, []).append(roll_no)
    
    # Store sorted roll numbers in the department-specific key in dict_dept
    dict_dept.setdefault(dept_id, {}).setdefault("all", []).extend(temp_list)

# Remove keys "gis" and "nan" from each year's dictionary
for year in dict_dept:
    dict_dept[year] = {k: v for k, v in dict_dept[year].items() if k not in ["gis", "nan"]}

print(dict_dept)



for x, y in output.items():
  print(x, y)

def allocation():
    temp_list = []
    currentIndex = 0
    for i in range(0,len(selectedHalls)):
        output.update({selectedHalls[i] : []})
#allocation()
print(output)